"""
make_all_figures.py — ONE file that regenerates EVERYTHING.

Run:  python3 make_all_figures.py

Folder layout (all relative to this "experiment 2" project root):

    experiment 2/
        code/                 <- this script lives here
        excel/                <- INPUT: VTC2_HUman.xlsx
        output/
            excel/            <- generated .xlsx + .csv
            figures/          <- generated .png + .pdf
                analysis2/    <- analysis2 stacked-bar / correlation figures

The pipeline runs in dependency order:

  1) 168-clip clean Excel + fig5 scatter
  2) figA/B/C/D detailed clip figures            (needs step 1)
  3) 14-recording Excel + figR_main/A/B/D        (needs step 1)
  4) experiment figure bundle PDF                 (needs steps 1-3)
  5) analysis2 correlations + stacked-bar figures
  6) analysis2 figure bundle PDF                  (needs step 5)

All paths are anchored to this file's location, so it can be launched from any
working directory.
"""

import os
import numpy as np
import pandas as pd
import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.image as mpimg
from matplotlib.lines import Line2D
from matplotlib.patches import Patch
from matplotlib.backends.backend_pdf import PdfPages
from scipy import stats

# ---------------------------------------------------------------- paths
CODE_DIR = os.path.dirname(os.path.abspath(__file__))      # .../experiment 2/code
ROOT     = os.path.dirname(CODE_DIR)                        # .../experiment 2
SRC      = os.path.join(ROOT, "excel", "VTC2_HUman.xlsx")  # INPUT excel

OUT      = os.path.join(ROOT, "output")
OUT_XL   = os.path.join(OUT, "excel")                      # generated .xlsx / .csv
OUT_FIG  = os.path.join(OUT, "figures")                    # generated .png / .pdf
OUT_A2   = os.path.join(OUT_FIG, "analysis2")              # analysis2 figures

CLEAN = os.path.join(OUT_XL, "VTC2_vs_Human_168clips.xlsx")  # produced by step 1

def XL(*parts):
    """Absolute path inside output/excel."""
    return os.path.join(OUT_XL, *parts)

def FIG(*parts):
    """Absolute path inside output/figures."""
    return os.path.join(OUT_FIG, *parts)

def A2(*parts):
    """Absolute path inside output/figures/analysis2."""
    return os.path.join(OUT_A2, *parts)

def ensure_dirs():
    for d in (OUT_XL, OUT_FIG, OUT_A2,
              A2("correlations"), A2("stacked_clip_together"), A2("stacked_clip_single"),
              A2("stacked_recording"), A2("stacked_permodel")):
        os.makedirs(d, exist_ok=True)

# panels shared by the experiment (VTC2 vs Human) figures
def experiment_panels(adult, kchi):
    return [
        ("Adult", adult, "Total",   "Adult — Total Words"),
        ("Adult", adult, "English", "Adult — English Words"),
        ("Adult", adult, "Spanish", "Adult — Spanish Words"),
        ("KCHI",  kchi,  "Total",   "Key Child — Total Words"),
        ("KCHI",  kchi,  "English", "Key Child — English Words"),
        ("KCHI",  kchi,  "Spanish", "Key Child — Spanish Words"),
    ]

def xy(df, metric):
    return (df[f"Hum_{metric}"].astype(float).values,
            df[f"VTC2_{metric}"].astype(float).values)


# ============================================================================
# STEP 1 — clean 168-clip Excel + fig5 scatter
# ============================================================================
def step1_168clips():
    print("\n=== STEP 1: 168-clip clean Excel + fig5 ===")
    HUM  = {"Total": 2, "English": 3, "Spanish": 4}      # Human Eval cols
    VTC2 = {"Total": 17, "English": 18, "Spanish": 19}   # VTC2 + Whisper cols

    def extract(sheet):
        wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))[3:]  # skip 3 header rows
        rec = None
        out = []
        for r in rows:
            if r[1] is None:
                continue
            if r[0] is not None:
                rec = r[0]
            clip = r[1]
            if str(clip).endswith("_COMBINED"):   # drop per-recording aggregate row
                continue
            if r[HUM["Total"]] is None:
                continue
            out.append({
                "Recording": rec, "Clip": clip,
                "Hum_Total": r[HUM["Total"]], "Hum_English": r[HUM["English"]], "Hum_Spanish": r[HUM["Spanish"]],
                "VTC2_Total": r[VTC2["Total"]], "VTC2_English": r[VTC2["English"]], "VTC2_Spanish": r[VTC2["Spanish"]],
            })
        return pd.DataFrame(out)

    def clean_na(df):
        num_cols = ["Hum_Total", "Hum_English", "Hum_Spanish",
                    "VTC2_Total", "VTC2_English", "VTC2_Spanish"]
        for c in num_cols:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
        return df

    adult = clean_na(extract("ADULT"))
    kchi = clean_na(extract("KCHI"))
    print("ADULT clips:", len(adult), "| KCHI clips:", len(kchi))
    print("Recordings:", adult["Recording"].nunique())

    with pd.ExcelWriter(CLEAN, engine="openpyxl") as xl:
        adult.to_excel(xl, sheet_name="ADULT", index=False)
        kchi.to_excel(xl, sheet_name="KCHI", index=False)
    print("Wrote excel/VTC2_vs_Human_168clips.xlsx")

    # ---- fig5: scatter colored by recording ----
    recs = sorted(adult["Recording"].unique())
    cmap = plt.get_cmap("tab20")
    colors = {rec: cmap(i % 20) for i, rec in enumerate(recs)}
    panels = experiment_panels(adult, kchi)

    fig, axes = plt.subplots(2, 3, figsize=(18, 12))
    axes = axes.ravel()
    for ax, (spk, df, metric, title) in zip(axes, panels):
        x = df[f"Hum_{metric}"].astype(float).values
        y = df[f"VTC2_{metric}"].astype(float).values
        for rec in recs:
            m = df["Recording"] == rec
            ax.scatter(df.loc[m, f"Hum_{metric}"], df.loc[m, f"VTC2_{metric}"],
                       color=colors[rec], s=55, alpha=0.85, edgecolors="white", linewidths=0.5)
        hi = max(x.max(), y.max()) * 1.05
        ax.plot([0, hi], [0, hi], color="gray", lw=1.2, label="Identity (y = x)")
        if len(x) > 1 and np.ptp(x) > 0:
            sl, ic, r, _, _ = stats.linregress(x, y)
            xs = np.array([0, hi])
            ax.plot(xs, sl * xs + ic, "k--", lw=1.4, label="OLS regression")
            ax.text(0.04, 0.94, f"r = {r:.2f}", transform=ax.transAxes,
                    fontsize=12, fontweight="bold", va="top",
                    bbox=dict(boxstyle="round", fc="white", ec="black"))
        ax.set_xlim(0, hi); ax.set_ylim(0, hi)
        ax.set_title(title, fontsize=13, fontweight="bold")
        ax.set_xlabel("Human Reference")
        ax.set_ylabel("VTC2 + WhisperX")
        ax.grid(True, ls="--", alpha=0.3)

    fig.suptitle("Per-Clip Word Counts: VTC2+WhisperX vs Human Reference\n"
                 "(N = 168 clips  •  14 recordings × 12 clips)",
                 fontsize=16, fontweight="bold")

    rec_handles = [Line2D([0], [0], marker="o", ls="", color=colors[r],
                          markersize=8, label=r) for r in recs]
    style_handles = [Line2D([0], [0], color="gray", lw=1.2, label="Identity (y = x)"),
                     Line2D([0], [0], color="black", ls="--", lw=1.4, label="OLS regression")]
    leg1 = fig.legend(handles=rec_handles, loc="lower center", ncol=7,
                      bbox_to_anchor=(0.5, -0.02), fontsize=9, title="Recording",
                      title_fontsize=10, frameon=True)
    fig.add_artist(leg1)
    fig.legend(handles=style_handles, loc="lower center", ncol=2,
               bbox_to_anchor=(0.5, -0.08), fontsize=10, frameon=False)

    fig.tight_layout(rect=[0, 0.06, 1, 0.95])
    fig.savefig(FIG("fig5_wordcount_scatter_168clips_bycolor.png"), dpi=150, bbox_inches="tight")
    plt.close(fig)
    print("Wrote figures/fig5_wordcount_scatter_168clips_bycolor.png")


# ============================================================================
# STEP 2 — figA/B/C/D detailed clip figures
# ============================================================================
def step2_more_figures():
    print("\n=== STEP 2: figA/B/C/D detailed clip figures ===")
    adult = pd.read_excel(CLEAN, sheet_name="ADULT")
    kchi  = pd.read_excel(CLEAN, sheet_name="KCHI")
    recs = sorted(adult["Recording"].unique())

    palette = plt.get_cmap("tab20").colors
    colors = {r: palette[i % 20] for i, r in enumerate(recs)}
    markers_cycle = ["o", "^", "s", "D", "P", "X", "v", "*", "<", ">", "h", "p", "d", "8"]
    markers = {r: markers_cycle[i % len(markers_cycle)] for i, r in enumerate(recs)}
    panels = experiment_panels(adult, kchi)

    # -------------------- A) scatter + stats
    fig, axes = plt.subplots(2, 3, figsize=(18, 12)); axes = axes.ravel()
    for ax, (spk, df, metric, title) in zip(axes, panels):
        x, y = xy(df, metric)
        for r in recs:
            m = df["Recording"] == r
            ax.scatter(df.loc[m, f"Hum_{metric}"], df.loc[m, f"VTC2_{metric}"],
                       color=colors[r], s=50, alpha=0.85, edgecolors="white", linewidths=0.4)
        hi = max(x.max(), y.max()) * 1.05
        ax.plot([0, hi], [0, hi], color="gray", lw=1.2)
        sl, ic, rr, _, _ = stats.linregress(x, y)
        xs = np.array([0, hi]); ax.plot(xs, sl*xs+ic, "k--", lw=1.4)
        mae = np.mean(np.abs(y - x)); bias = np.mean(y - x)
        txt = (f"r = {rr:.2f}   R² = {rr**2:.2f}\n"
               f"slope = {sl:.2f}   int = {ic:.0f}\n"
               f"MAE = {mae:.0f} words\n"
               f"bias = {bias:+.0f} (VTC2−Human)")
        ax.text(0.04, 0.96, txt, transform=ax.transAxes, fontsize=10.5, va="top",
                bbox=dict(boxstyle="round", fc="white", ec="black", alpha=0.9))
        ax.set_xlim(0, hi); ax.set_ylim(0, hi)
        ax.set_title(title, fontsize=13, fontweight="bold")
        ax.set_xlabel("Human Reference"); ax.set_ylabel("VTC2 + WhisperX")
        ax.grid(True, ls="--", alpha=0.3)
    fig.suptitle("VTC2+WhisperX vs Human — Per-Clip with Regression Stats "
                 "(N = 168 clips)", fontsize=16, fontweight="bold")
    rh = [Line2D([0],[0], marker="o", ls="", color=colors[r], markersize=8, label=r) for r in recs]
    fig.legend(handles=rh, loc="lower center", ncol=7, bbox_to_anchor=(0.5, -0.02),
               fontsize=9, title="Recording", frameon=True)
    fig.tight_layout(rect=[0, 0.05, 1, 0.95])
    fig.savefig(FIG("figA_scatter_with_stats.png"), dpi=150, bbox_inches="tight")
    plt.close(fig)
    print("Wrote figures/figA_scatter_with_stats.png")

    # -------------------- B) Bland-Altman
    fig, axes = plt.subplots(2, 3, figsize=(18, 12)); axes = axes.ravel()
    for ax, (spk, df, metric, title) in zip(axes, panels):
        x, y = xy(df, metric)
        mean = (x + y) / 2.0; diff = y - x
        bias = diff.mean(); sd = diff.std(ddof=1)
        lo, up = bias - 1.96*sd, bias + 1.96*sd
        for r in recs:
            m = (df["Recording"] == r).values
            ax.scatter(mean[m], diff[m], color=colors[r], s=50, alpha=0.85,
                       edgecolors="white", linewidths=0.4)
        ax.axhline(bias, color="k", ls="-", lw=1.4)
        ax.axhline(up, color="r", ls="--", lw=1.2); ax.axhline(lo, color="r", ls="--", lw=1.2)
        ax.axhline(0, color="gray", ls=":", lw=1.0)
        xmax = mean.max()*1.05
        ax.text(xmax, bias, f" bias={bias:+.0f}", va="center", fontsize=9, color="k")
        ax.text(xmax, up, f" +1.96SD={up:+.0f}", va="center", fontsize=8, color="r")
        ax.text(xmax, lo, f" −1.96SD={lo:+.0f}", va="center", fontsize=8, color="r")
        ax.set_title(title, fontsize=13, fontweight="bold")
        ax.set_xlabel("Mean of Human & VTC2  ((x+y)/2)")
        ax.set_ylabel("VTC2 − Human")
        ax.grid(True, ls="--", alpha=0.3)
    fig.suptitle("Bland–Altman Agreement: VTC2+WhisperX vs Human "
                 "(negative = VTC2 under-counts)", fontsize=16, fontweight="bold")
    fig.legend(handles=rh, loc="lower center", ncol=7, bbox_to_anchor=(0.5, -0.02),
               fontsize=9, title="Recording", frameon=True)
    fig.tight_layout(rect=[0, 0.05, 1, 0.95])
    fig.savefig(FIG("figB_bland_altman.png"), dpi=150, bbox_inches="tight")
    plt.close(fig)
    print("Wrote figures/figB_bland_altman.png")

    # -------------------- C) per-recording small multiples
    fig, axes = plt.subplots(3, 5, figsize=(20, 12)); axes = axes.ravel()
    for ax, r in zip(axes, recs):
        a = adult[adult["Recording"] == r]; k = kchi[kchi["Recording"] == r]
        ax.scatter(a["Hum_Total"], a["VTC2_Total"], color="#1f77b4", s=45,
                   alpha=0.85, label="Adult", edgecolors="white", linewidths=0.4)
        ax.scatter(k["Hum_Total"], k["VTC2_Total"], color="#d62728", s=45,
                   alpha=0.85, label="Key Child", edgecolors="white", linewidths=0.4)
        allx = pd.concat([a["Hum_Total"], k["Hum_Total"], a["VTC2_Total"], k["VTC2_Total"]]).astype(float)
        hi = allx.max()*1.08 if allx.max() > 0 else 1
        ax.plot([0, hi], [0, hi], color="gray", lw=1.0)
        r_ad = stats.pearsonr(a["Hum_Total"].astype(float), a["VTC2_Total"].astype(float))[0]
        r_kc = stats.pearsonr(k["Hum_Total"].astype(float), k["VTC2_Total"].astype(float))[0]
        ax.set_xlim(0, hi); ax.set_ylim(0, hi)
        ax.set_title(f"{r}\nAdult r={r_ad:.2f}   Key Child r={r_kc:.2f}",
                     fontsize=9.5, fontweight="bold")
        ax.grid(True, ls="--", alpha=0.3)
    for ax in axes[len(recs):]:
        ax.axis("off")
    axes[0].legend(loc="upper left", fontsize=9)
    fig.suptitle("Per-Recording Agreement — Total Words (each panel = 1 recording × 12 clips)",
                 fontsize=16, fontweight="bold")
    fig.supxlabel("Human Reference"); fig.supylabel("VTC2 + WhisperX")
    fig.tight_layout(rect=[0, 0.02, 1, 0.95])
    fig.savefig(FIG("figC_per_recording_smallmultiples.png"), dpi=150, bbox_inches="tight")
    plt.close(fig)
    print("Wrote figures/figC_per_recording_smallmultiples.png")

    # -------------------- D) color + shape scatter
    fig, axes = plt.subplots(2, 3, figsize=(18, 12)); axes = axes.ravel()
    for ax, (spk, df, metric, title) in zip(axes, panels):
        x, y = xy(df, metric)
        for r in recs:
            m = df["Recording"] == r
            ax.scatter(df.loc[m, f"Hum_{metric}"], df.loc[m, f"VTC2_{metric}"],
                       color=colors[r], marker=markers[r], s=60, alpha=0.9,
                       edgecolors="black", linewidths=0.4)
        hi = max(x.max(), y.max())*1.05
        ax.plot([0, hi], [0, hi], color="gray", lw=1.2)
        sl, ic, rr, _, _ = stats.linregress(x, y)
        xs = np.array([0, hi]); ax.plot(xs, sl*xs+ic, "k--", lw=1.4)
        ax.text(0.04, 0.94, f"r = {rr:.2f}", transform=ax.transAxes, fontsize=12,
                fontweight="bold", va="top",
                bbox=dict(boxstyle="round", fc="white", ec="black"))
        ax.set_xlim(0, hi); ax.set_ylim(0, hi)
        ax.set_title(title, fontsize=13, fontweight="bold")
        ax.set_xlabel("Human Reference"); ax.set_ylabel("VTC2 + WhisperX")
        ax.grid(True, ls="--", alpha=0.3)
    fig.suptitle("VTC2+WhisperX vs Human — Distinct Color + Shape per Recording "
                 "(N = 168 clips)", fontsize=16, fontweight="bold")
    rh2 = [Line2D([0],[0], marker=markers[r], ls="", color=colors[r], markeredgecolor="black",
                  markersize=9, label=r) for r in recs]
    fig.legend(handles=rh2, loc="lower center", ncol=7, bbox_to_anchor=(0.5, -0.02),
               fontsize=9, title="Recording", frameon=True)
    fig.tight_layout(rect=[0, 0.05, 1, 0.95])
    fig.savefig(FIG("figD_scatter_color_shape.png"), dpi=150, bbox_inches="tight")
    plt.close(fig)
    print("Wrote figures/figD_scatter_color_shape.png")


# ============================================================================
# STEP 3 — 14-recording Excel + figR_*
# ============================================================================
def step3_recording_level():
    print("\n=== STEP 3: 14-recording Excel + figR figures ===")
    cols = ["Hum_Total","Hum_English","Hum_Spanish","VTC2_Total","VTC2_English","VTC2_Spanish"]

    def agg(sheet):
        df = pd.read_excel(CLEAN, sheet_name=sheet)
        return df.groupby("Recording", as_index=False)[cols].sum()

    adult = agg("ADULT"); kchi = agg("KCHI")
    recs = sorted(adult["Recording"].unique())
    palette = plt.get_cmap("tab20").colors
    colors = {r: palette[i % 20] for i, r in enumerate(recs)}
    mk = ["o","^","s","D","P","X","v","*","<",">","h","p","d","8"]
    markers = {r: mk[i % len(mk)] for i, r in enumerate(recs)}

    with pd.ExcelWriter(XL("VTC2_vs_Human_14recordings.xlsx"), engine="openpyxl") as xl:
        adult.to_excel(xl, sheet_name="ADULT", index=False)
        kchi.to_excel(xl, sheet_name="KCHI", index=False)
    print("Wrote excel/VTC2_vs_Human_14recordings.xlsx  (N=14 per speaker)")

    panels = experiment_panels(adult, kchi)
    rh  = [Line2D([0],[0],marker="o",ls="",color=colors[r],markersize=8,label=r) for r in recs]
    rh2 = [Line2D([0],[0],marker=markers[r],ls="",color=colors[r],markeredgecolor="black",
                  markersize=9,label=r) for r in recs]

    def scatter_fig(with_stats, with_shape, fname, suptitle):
        fig, axes = plt.subplots(2, 3, figsize=(18, 12)); axes = axes.ravel()
        for ax,(spk,df,m,title) in zip(axes,panels):
            x,y = xy(df,m)
            for r in recs:
                sub = df[df["Recording"]==r]
                ax.scatter(sub[f"Hum_{m}"], sub[f"VTC2_{m}"], color=colors[r],
                           marker=markers[r] if with_shape else "o", s=110, alpha=0.9,
                           edgecolors="black", linewidths=0.5)
            hi = max(x.max(), y.max())*1.05
            ax.plot([0,hi],[0,hi],color="gray",lw=1.2)
            sl,ic,rr,_,_ = stats.linregress(x,y)
            xs=np.array([0,hi]); ax.plot(xs,sl*xs+ic,"k--",lw=1.4)
            if with_stats:
                mae=np.mean(np.abs(y-x)); bias=np.mean(y-x)
                ax.text(0.04,0.96,f"r = {rr:.2f}   R² = {rr**2:.2f}\nslope = {sl:.2f}   int = {ic:.0f}\n"
                        f"MAE = {mae:.0f} words\nbias = {bias:+.0f} (VTC2−Human)",
                        transform=ax.transAxes,fontsize=10.5,va="top",
                        bbox=dict(boxstyle="round",fc="white",ec="black",alpha=0.9))
            else:
                ax.text(0.04,0.94,f"r = {rr:.2f}",transform=ax.transAxes,fontsize=12,
                        fontweight="bold",va="top",bbox=dict(boxstyle="round",fc="white",ec="black"))
            ax.set_xlim(0,hi); ax.set_ylim(0,hi)
            ax.set_title(title,fontsize=13,fontweight="bold")
            ax.set_xlabel("Human Reference"); ax.set_ylabel("VTC2 + WhisperX")
            ax.grid(True,ls="--",alpha=0.3)
        fig.suptitle(suptitle,fontsize=16,fontweight="bold")
        fig.legend(handles=(rh2 if with_shape else rh),loc="lower center",ncol=7,
                   bbox_to_anchor=(0.5,-0.02),fontsize=9,title="Recording",frameon=True)
        fig.tight_layout(rect=[0,0.05,1,0.95])
        fig.savefig(FIG(fname),dpi=150,bbox_inches="tight"); plt.close(fig)
        print("Wrote figures/"+fname)

    scatter_fig(False, False, "figR_main_scatter.png",
                "VTC2+WhisperX vs Human — Per-Recording (N = 14 recordings)")
    scatter_fig(True, False, "figR_A_scatter_with_stats.png",
                "VTC2+WhisperX vs Human — Per-Recording with Stats (N = 14)")
    scatter_fig(False, True, "figR_D_color_shape.png",
                "VTC2+WhisperX vs Human — Per-Recording, Color + Shape (N = 14)")

    # Bland-Altman (N=14)
    fig, axes = plt.subplots(2,3,figsize=(18,12)); axes=axes.ravel()
    for ax,(spk,df,m,title) in zip(axes,panels):
        x,y = xy(df,m); mean=(x+y)/2; diff=y-x
        bias=diff.mean(); sd=diff.std(ddof=1); lo,up=bias-1.96*sd,bias+1.96*sd
        for r in recs:
            mm=(df["Recording"]==r).values
            ax.scatter(mean[mm],diff[mm],color=colors[r],s=110,alpha=0.9,
                       edgecolors="black",linewidths=0.5)
        ax.axhline(bias,color="k",lw=1.4); ax.axhline(up,color="r",ls="--",lw=1.2)
        ax.axhline(lo,color="r",ls="--",lw=1.2); ax.axhline(0,color="gray",ls=":",lw=1.0)
        xmax=mean.max()*1.05
        ax.text(xmax,bias,f" bias={bias:+.0f}",va="center",fontsize=9)
        ax.text(xmax,up,f" +1.96SD={up:+.0f}",va="center",fontsize=8,color="r")
        ax.text(xmax,lo,f" −1.96SD={lo:+.0f}",va="center",fontsize=8,color="r")
        ax.set_title(title,fontsize=13,fontweight="bold")
        ax.set_xlabel("Mean of Human & VTC2  ((x+y)/2)"); ax.set_ylabel("VTC2 − Human")
        ax.grid(True,ls="--",alpha=0.3)
    fig.suptitle("Bland–Altman Agreement — Per-Recording (N = 14)",fontsize=16,fontweight="bold")
    fig.legend(handles=rh,loc="lower center",ncol=7,bbox_to_anchor=(0.5,-0.02),
               fontsize=9,title="Recording",frameon=True)
    fig.tight_layout(rect=[0,0.05,1,0.95])
    fig.savefig(FIG("figR_B_bland_altman.png"),dpi=150,bbox_inches="tight"); plt.close(fig)
    print("Wrote figures/figR_B_bland_altman.png")

    print("Per-recording r values:")
    for spk,df,m,_ in panels:
        x,y=xy(df,m); print(f"  {spk:5s} {m:8s} r={stats.pearsonr(x,y)[0]:.2f}")


# ============================================================================
# STEP 4 — experiment figure bundle PDF
# ============================================================================
def step4_experiment_pdf():
    print("\n=== STEP 4: experiment figure bundle PDF ===")
    PAGES = [
        ("SECTION", "Clip-Level Analysis  (N = 168 clips = 14 recordings x 12 clips)"),
        ("fig5_wordcount_scatter_168clips_bycolor.png",
         "Per-clip scatter: VTC2+WhisperX vs Human, 14 recordings by color (12 dots each)."),
        ("figA_scatter_with_stats.png",
         "Per-clip scatter with regression stats (r, R^2, slope, intercept, MAE, bias)."),
        ("figB_bland_altman.png",
         "Bland-Altman agreement (clip level): difference vs mean, with bias +/-1.96 SD limits."),
        ("figC_per_recording_smallmultiples.png",
         "Per-recording small multiples (Total Words): each panel = 1 recording, Adult & Key Child r shown."),
        ("figD_scatter_color_shape.png",
         "Per-clip scatter with distinct color + marker shape per recording."),
        ("SECTION", "Recording-Level Analysis  (N = 14 recordings; each = sum of its 12 clips)"),
        ("figR_main_scatter.png",
         "Per-recording scatter: VTC2+WhisperX vs Human, 14 points."),
        ("figR_A_scatter_with_stats.png",
         "Per-recording scatter with regression stats."),
        ("figR_B_bland_altman.png",
         "Bland-Altman agreement (recording level)."),
        ("figR_D_color_shape.png",
         "Per-recording scatter with distinct color + marker shape."),
    ]
    OUT_PDF = FIG("VTC2_vs_Human_figures_bundle.pdf")
    with PdfPages(OUT_PDF) as pdf:
        fig = plt.figure(figsize=(11, 8.5)); fig.subplots_adjust(0,0,1,1)
        ax = fig.add_axes([0,0,1,1]); ax.axis("off")
        ax.text(0.5, 0.62, "VTC2 + WhisperX vs Human Reference", ha="center",
                fontsize=24, fontweight="bold")
        ax.text(0.5, 0.54, "Word-Count Agreement — Figure Bundle", ha="center", fontsize=16)
        ax.text(0.5, 0.44,
                "168 clips (14 recordings x 12 clips)  +  14 recording-level summaries\n"
                "Speakers: Adult & Key Child   Metrics: Total / English / Spanish words\n"
                "N/A treated as 0 (silent clips where Human = 0 too)",
                ha="center", fontsize=12, linespacing=1.6)
        pdf.savefig(fig); plt.close(fig)

        for item, caption in PAGES:
            if item == "SECTION":
                fig = plt.figure(figsize=(11, 8.5)); ax = fig.add_axes([0,0,1,1]); ax.axis("off")
                ax.text(0.5, 0.5, caption, ha="center", va="center",
                        fontsize=20, fontweight="bold", wrap=True)
                pdf.savefig(fig); plt.close(fig); continue
            img = mpimg.imread(FIG(item))
            fig = plt.figure(figsize=(11, 8.5))
            ax = fig.add_axes([0.02, 0.06, 0.96, 0.88]); ax.axis("off")
            ax.imshow(img)
            fig.text(0.5, 0.025, caption, ha="center", fontsize=9, wrap=True)
            pdf.savefig(fig, dpi=200); plt.close(fig)
    print("Wrote figures/VTC2_vs_Human_figures_bundle.pdf")


# ============================================================================
# STEP 5 — analysis2 correlations + stacked bars
# ============================================================================
def step5_analysis2():
    print("\n=== STEP 5: analysis2 correlations + stacked bars ===")
    SHEET = "OVERALL"
    MODELS = {
        "Human Eval":          (2, 3, 4),
        "Human-TS + Whisper":  (7, 8, 9),
        "VTC1 + Whisper":      (12, 13, 14),
        "VTC2 + Whisper":      (17, 18, 19),
        "Only Whisper":        (22, 23, 24),
    }
    SHORT = {"Human Eval": "Human", "Human-TS + Whisper": "HumanTS",
             "VTC1 + Whisper": "VTC1", "VTC2 + Whisper": "VTC2", "Only Whisper": "OnlyWhisper"}
    ENG_C, SPA_C = "#3b76af", "#e1812c"

    def num(v): return 0.0 if v in ("N/A", None) else float(v)

    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))[3:]
    rec = None; records = []
    for r in rows:
        if r[1] is None: continue
        if r[0] is not None: rec = r[0]
        if str(r[1]).endswith("_COMBINED"): continue
        if r[MODELS["Human Eval"][0]] is None: continue
        d = {"Recording": rec, "Clip": r[1]}
        for m, (t, e, s) in MODELS.items():
            d[f"{m}|Eng"] = num(r[e]); d[f"{m}|Spa"] = num(r[s])
        records.append(d)
    clip = pd.DataFrame(records)
    clip["ClipNo"] = clip["Clip"].str.extract(r"_(\d+)$")[0]
    recs = list(dict.fromkeys(clip["Recording"]))
    print(f"clips={len(clip)}  recordings={len(recs)}  per-rec={clip.groupby('Recording').size().unique()}")

    val_cols = [c for c in clip.columns if "|" in c]
    recdf = clip.groupby("Recording", as_index=False)[val_cols].sum()

    with pd.ExcelWriter(XL("analysis2_data_tidy.xlsx"), engine="openpyxl") as xl:
        clip.to_excel(xl, sheet_name="clip_level", index=False)
        recdf.to_excel(xl, sheet_name="recording_level", index=False)
    print("Wrote excel/analysis2_data_tidy.xlsx")

    # ---------------- correlations ----------------
    def corr_table(df, level):
        out = []
        for m in MODELS:
            row = {"Model": m, "Level": level}
            for lang in ["Eng", "Spa"]:
                x = df[f"Human Eval|{lang}"].values.astype(float)
                y = df[f"{m}|{lang}"].values.astype(float)
                if np.std(x) == 0 or np.std(y) == 0:
                    row[f"{lang}_r"] = np.nan
                else:
                    row[f"{lang}_r"] = stats.pearsonr(x, y)[0]
            out.append(row)
        return pd.DataFrame(out)

    ct_clip = corr_table(clip, "clip")
    ct_rec = corr_table(recdf, "recording")
    corr_all = pd.concat([ct_clip, ct_rec], ignore_index=True)
    corr_all.to_csv(XL("correlations_model_vs_human.csv"), index=False)
    print("Wrote excel/correlations_model_vs_human.csv")
    print("Correlations (model tokens vs Human tokens):")
    print(corr_all.to_string(index=False))

    def heatmap(ct, level, fname):
        mods = ct["Model"].tolist()
        data = ct[["Eng_r", "Spa_r"]].values.astype(float)
        fig, ax = plt.subplots(figsize=(6, 5))
        im = ax.imshow(data, cmap="RdYlGn", vmin=0, vmax=1, aspect="auto")
        ax.set_xticks([0, 1]); ax.set_xticklabels(["English tokens", "Spanish tokens"])
        ax.set_yticks(range(len(mods))); ax.set_yticklabels(mods)
        for i in range(len(mods)):
            for j in range(2):
                v = data[i, j]
                ax.text(j, i, "1.00 (self)" if (mods[i] == "Human Eval") else f"{v:.2f}",
                        ha="center", va="center", fontsize=10,
                        color="black", fontweight="bold")
        ax.set_title(f"Correlation with Human tokens — {level} level\n(Pearson r, per language)",
                     fontsize=12, fontweight="bold")
        fig.colorbar(im, ax=ax, label="Pearson r")
        fig.tight_layout(); fig.savefig(fname, dpi=150, bbox_inches="tight"); plt.close(fig)

    heatmap(ct_clip, "Clip (N=168)", A2("correlations", "heatmap_clip.png"))
    heatmap(ct_rec, "Recording (N=14)", A2("correlations", "heatmap_recording.png"))

    # ---------------- stacked bars ----------------
    LEG = [Patch(fc=ENG_C, label="English tokens"), Patch(fc=SPA_C, label="Spanish tokens")]
    model_list = list(MODELS.keys())
    ORDER_NOTE = "Bars per group (L→R): " + ", ".join(SHORT[m] for m in model_list)

    HATCH = {m: h for m, h in zip(model_list, ["", "///", "...", "xxx", "\\\\\\"])}
    CODE  = {"Human Eval": "H", "Human-TS + Whisper": "HT", "VTC1 + Whisper": "V1",
             "VTC2 + Whisper": "V2", "Only Whisper": "OW"}
    MODEL_LEG = [Patch(fc="0.85", ec="black", hatch=HATCH[m], label=f"{CODE[m]} = {m}")
                 for m in model_list]

    def stack(ax, xpos, eng, spa, width, hatch=None):
        ax.bar(xpos, eng, width, color=ENG_C, edgecolor="black", linewidth=0.3, hatch=hatch)
        ax.bar(xpos, spa, width, bottom=eng, color=SPA_C, edgecolor="black", linewidth=0.3, hatch=hatch)

    # clip-level TOGETHER: 14 figs
    GROUP = 3.0
    for rname in recs:
        sub = clip[clip["Recording"] == rname].reset_index(drop=True)
        n = len(sub); w = 0.34
        centers = np.arange(n) * GROUP
        fig, ax = plt.subplots(figsize=(26, 8))
        for j, m in enumerate(model_list):
            off = (j - (len(model_list)-1)/2) * w
            stack(ax, centers + off, sub[f"{m}|Eng"].values, sub[f"{m}|Spa"].values, w, hatch=HATCH[m])
        trans = ax.get_xaxis_transform()
        for i in range(n):
            for j, m in enumerate(model_list):
                off = (j - (len(model_list)-1)/2) * w
                ax.text(centers[i] + off, -0.015, CODE[m], transform=trans,
                        ha="center", va="top", fontsize=7, rotation=90)
        for i in range(n):
            ax.text(centers[i], -0.11, str(sub["ClipNo"].iloc[i]), transform=trans,
                    ha="center", va="top", fontsize=12, fontweight="bold")
        ax.text(0.5, -0.17, "Clip", transform=ax.transAxes, ha="center", va="top", fontsize=12)
        ax.set_xticks([])
        ax.set_ylim(bottom=0); ax.set_xlim(centers[0]-GROUP/2, centers[-1]+GROUP/2)
        ax.set_ylabel("Tokens (words)", fontsize=12)
        ax.set_title(f"English vs Spanish tokens per clip — {rname}\n{ORDER_NOTE}",
                     fontsize=13, fontweight="bold")
        ax.grid(True, axis="y", ls="--", alpha=0.3)
        ax.legend(handles=LEG + MODEL_LEG, loc="upper center", bbox_to_anchor=(0.5, -0.22),
                  ncol=7, frameon=True, fontsize=10)
        fig.tight_layout()
        fig.savefig(A2("stacked_clip_together", f"{rname}.png"), dpi=130, bbox_inches="tight")
        plt.close(fig)
    print(f"Wrote {len(recs)} clip-level TOGETHER figures")

    # clip-level SINGLE: one fig per model
    for m in model_list:
        fig, axes = plt.subplots(3, 5, figsize=(22, 12)); axes = axes.ravel()
        for ax, rname in zip(axes, recs):
            sub = clip[clip["Recording"] == rname].reset_index(drop=True)
            n = len(sub)
            stack(ax, np.arange(n), sub[f"{m}|Eng"].values, sub[f"{m}|Spa"].values, 0.8)
            ax.set_xticks(range(n)); ax.set_xticklabels(sub["ClipNo"], fontsize=7, rotation=90)
            ax.set_title(rname, fontsize=9, fontweight="bold")
            ax.grid(True, axis="y", ls="--", alpha=0.3)
        for ax in axes[len(recs):]: ax.axis("off")
        axes[0].legend(handles=LEG, loc="upper left", fontsize=8)
        fig.suptitle(f"English vs Spanish tokens per clip — Model: {m}  "
                     f"(each panel = 1 recording x 12 clips)", fontsize=15, fontweight="bold")
        fig.supxlabel("Clip"); fig.supylabel("Tokens (words)")
        fig.tight_layout(rect=[0, 0.01, 1, 0.96])
        fig.savefig(A2("stacked_clip_single", f"{SHORT[m]}.png"), dpi=130, bbox_inches="tight")
        plt.close(fig)
    print(f"Wrote {len(model_list)} clip-level SINGLE figures (one per model)")

    # recording-level TOGETHER
    RGROUP = 1.6
    n = len(recs); w = 0.26
    rcent = np.arange(n) * RGROUP
    fig, ax = plt.subplots(figsize=(22, 8))
    for j, m in enumerate(model_list):
        off = (j - (len(model_list)-1)/2) * w
        stack(ax, rcent + off, recdf[f"{m}|Eng"].values, recdf[f"{m}|Spa"].values, w, hatch=HATCH[m])
    rtrans = ax.get_xaxis_transform()
    for i in range(n):
        for j, m in enumerate(model_list):
            off = (j - (len(model_list)-1)/2) * w
            ax.text(rcent[i] + off, -0.015, CODE[m], transform=rtrans,
                    ha="center", va="top", fontsize=7, rotation=90)
    for i in range(n):
        label = recdf["Recording"].iloc[i].replace("_pre", "")
        ax.text(rcent[i], -0.11, label, transform=rtrans,
                ha="center", va="top", fontsize=9, rotation=30, rotation_mode="anchor")
    ax.set_xticks([])
    ax.set_xlim(rcent[0]-RGROUP/2, rcent[-1]+RGROUP/2); ax.set_ylim(bottom=0)
    ax.set_ylabel("Tokens (words)", fontsize=12)
    ax.set_title(f"English vs Spanish tokens per recording (N=14)\n{ORDER_NOTE}",
                 fontsize=13, fontweight="bold")
    ax.grid(True, axis="y", ls="--", alpha=0.3)
    ax.legend(handles=LEG + MODEL_LEG, loc="upper center", bbox_to_anchor=(0.5, -0.20),
              ncol=7, frameon=True, fontsize=10)
    fig.tight_layout(); fig.savefig(A2("stacked_recording", "recording_together.png"), dpi=140, bbox_inches="tight")
    plt.close(fig)

    # recording-level SINGLE
    for m in model_list:
        fig, ax = plt.subplots(figsize=(12, 6))
        stack(ax, np.arange(n), recdf[f"{m}|Eng"].values, recdf[f"{m}|Spa"].values, 0.7)
        ax.set_xticks(range(n)); ax.set_xticklabels(recdf["Recording"], rotation=45, ha="right", fontsize=8)
        ax.set_ylabel("Tokens (words)")
        ax.set_title(f"English vs Spanish tokens per recording — Model: {m}",
                     fontsize=13, fontweight="bold")
        ax.legend(handles=LEG, loc="upper right"); ax.grid(True, axis="y", ls="--", alpha=0.3)
        fig.tight_layout()
        fig.savefig(A2("stacked_recording", f"recording_single_{SHORT[m]}.png"), dpi=140, bbox_inches="tight")
        plt.close(fig)
    print("Wrote recording-level TOGETHER + 5 SINGLE figures")

    # per-model SUMMARY
    tot = {m: (clip[f"{m}|Eng"].sum(), clip[f"{m}|Spa"].sum()) for m in model_list}
    fig, ax = plt.subplots(figsize=(9, 6))
    xs = np.arange(len(model_list))
    eng = [tot[m][0] for m in model_list]; spa = [tot[m][1] for m in model_list]
    ax.bar(xs, eng, 0.6, color=ENG_C, label="English tokens")
    ax.bar(xs, spa, 0.6, bottom=eng, color=SPA_C, label="Spanish tokens")
    for i, m in enumerate(model_list):
        ax.text(i, eng[i]+spa[i], f"{int(eng[i]+spa[i])}", ha="center", va="bottom", fontsize=9)
    ax.set_xticks(xs); ax.set_xticklabels([SHORT[m] for m in model_list])
    ax.set_ylabel("Total tokens (all 168 clips)")
    ax.set_title("Total English vs Spanish tokens by model (OVERALL)", fontsize=13, fontweight="bold")
    ax.legend(); ax.grid(True, axis="y", ls="--", alpha=0.3)
    fig.tight_layout(); fig.savefig(A2("stacked_permodel", "summary_all_models.png"), dpi=150, bbox_inches="tight")
    plt.close(fig)
    for m in model_list:
        fig, ax = plt.subplots(figsize=(3.2, 5))
        e, s = tot[m]
        ax.bar([0], [e], 0.6, color=ENG_C, label="English")
        ax.bar([0], [s], 0.6, bottom=[e], color=SPA_C, label="Spanish")
        ax.text(0, e/2, f"{int(e)}", ha="center", va="center", color="white", fontsize=10, fontweight="bold")
        ax.text(0, e+s/2, f"{int(s)}", ha="center", va="center", color="white", fontsize=10, fontweight="bold")
        ax.set_xticks([0]); ax.set_xticklabels([SHORT[m]]); ax.set_ylabel("Total tokens")
        ax.set_title(m, fontsize=11, fontweight="bold"); ax.legend(fontsize=8)
        fig.tight_layout(); fig.savefig(A2("stacked_permodel", f"single_{SHORT[m]}.png"), dpi=150, bbox_inches="tight")
        plt.close(fig)
    print("Wrote per-model summary + 5 individual figures")


# ============================================================================
# STEP 6 — analysis2 figure bundle PDF
# ============================================================================
def step6_analysis2_pdf():
    print("\n=== STEP 6: analysis2 figure bundle PDF ===")
    MODELS_SHORT = ["Human", "HumanTS", "VTC1", "VTC2", "OnlyWhisper"]
    RECS = [
        "DL-0122001008_pre","DL-1223001005_pre","DL-1422001003_pre","DL-1522001002_pre",
        "DL-1523001001_pre","DL-1722001001_pre","DL-3023001001_pre","DL-3224001001_pre",
        "DL-3323002002_pre","DL-3423001006_pre","DL-3623001003_pre","DL-3923001002_pre",
        "DL-4023001006_pre","DL-5624002008_pre",
    ]
    PAGES = [("SECTION", "1. Correlations — model tokens vs Human (English & Spanish)")]
    PAGES += [("correlations/heatmap_clip.png", "Correlation with Human tokens — clip level (N=168)."),
              ("correlations/heatmap_recording.png", "Correlation with Human tokens — recording level (N=14).")]
    PAGES += [("SECTION", "2. Per-model token totals (English vs Spanish, OVERALL)")]
    PAGES += [("stacked_permodel/summary_all_models.png", "Total English vs Spanish tokens by model (all 168 clips).")]
    PAGES += [(f"stacked_permodel/single_{s}.png", f"Per-model total — {s}.") for s in MODELS_SHORT]
    PAGES += [("SECTION", "3. Recording-level stacked bars (N=14)")]
    PAGES += [("stacked_recording/recording_together.png", "All 5 models grouped per recording.")]
    PAGES += [(f"stacked_recording/recording_single_{s}.png", f"Recording level — {s}.") for s in MODELS_SHORT]
    PAGES += [("SECTION", "4. Clip-level stacked bars — SINGLE (one page per model, 14 recordings each)")]
    PAGES += [(f"stacked_clip_single/{s}.png", f"Clip level, single model — {s} (14 recordings x 12 clips).") for s in MODELS_SHORT]
    PAGES += [("SECTION", "5. Clip-level stacked bars — TOGETHER (one page per recording, 5 models grouped)")]
    PAGES += [(f"stacked_clip_together/{r}.png", f"Clip level, all models — {r} (12 clips x 5 models).") for r in RECS]

    OUT_PDF = FIG("analysis2_figures_bundle.pdf")
    n_fig = sum(1 for it, _ in PAGES if it != "SECTION")
    with PdfPages(OUT_PDF) as pdf:
        fig = plt.figure(figsize=(11, 8.5)); ax = fig.add_axes([0,0,1,1]); ax.axis("off")
        ax.text(0.5, 0.60, "analysis2 — English vs Spanish Token Analysis", ha="center",
                fontsize=22, fontweight="bold")
        ax.text(0.5, 0.52, "5 models vs Human  •  Speaker = OVERALL", ha="center", fontsize=15)
        ax.text(0.5, 0.42,
                "Models: Human Eval, Human-TS+Whisper, VTC1+Whisper, VTC2+Whisper, Only Whisper\n"
                "168 clips (14 recordings x 12 clips)   •   tokens = English/Spanish word counts\n"
                "Correlations + stacked bars (clip, recording, per-model; single & together)",
                ha="center", fontsize=11, linespacing=1.7)
        pdf.savefig(fig); plt.close(fig)

        for item, caption in PAGES:
            if item == "SECTION":
                fig = plt.figure(figsize=(11, 8.5)); ax = fig.add_axes([0,0,1,1]); ax.axis("off")
                ax.text(0.5, 0.5, caption, ha="center", va="center",
                        fontsize=18, fontweight="bold", wrap=True)
                pdf.savefig(fig); plt.close(fig); continue
            path = A2(*item.split("/"))
            if not os.path.exists(path):
                print("MISSING:", item); continue
            img = mpimg.imread(path); h, w = img.shape[:2]
            wide = (w / h) > 1.6
            fig = plt.figure(figsize=(14, 8.5) if wide else (11, 8.5))
            ax = fig.add_axes([0.02, 0.06, 0.96, 0.88]); ax.axis("off"); ax.imshow(img)
            fig.text(0.5, 0.02, caption, ha="center", fontsize=9)
            pdf.savefig(fig, dpi=200); plt.close(fig)
    print(f"Wrote figures/analysis2_figures_bundle.pdf  ({n_fig} figures + cover + section dividers)")


# ============================================================================
def main():
    print("Regenerating ALL figures from", SRC)
    ensure_dirs()
    step1_168clips()          # -> excel/VTC2_vs_Human_168clips.xlsx + fig5
    step2_more_figures()      # -> figA/B/C/D          (needs step 1)
    step3_recording_level()   # -> 14-recording xlsx + figR_*  (needs step 1)
    step4_experiment_pdf()    # -> experiment PDF      (needs steps 1-3)
    step5_analysis2()         # -> analysis2 figures
    step6_analysis2_pdf()     # -> analysis2 PDF       (needs step 5)
    print("\nALL DONE — every figure, Excel and PDF written under output/.")


if __name__ == "__main__":
    main()
